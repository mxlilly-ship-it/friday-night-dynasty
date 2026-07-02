#!/usr/bin/env python3
"""Build systems/playbook_overlay.py from FND_Playbook_Reference.xlsx.

Usage:
  python scripts/build_playbook_overlay.py [path/to/FND_Playbook_Reference.xlsx]
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

DEFAULT_XLSX = Path(r"C:\Users\mxlil\Desktop\Screenshots\FND_Playbook_Reference.xlsx")
OUT_PATH = ROOT / "systems" / "playbook_overlay.py"

OFF_CAT = {
    "Inside Run": "INSIDE_RUN",
    "Outside Run": "OUTSIDE_RUN",
    "Short Pass": "SHORT_PASS",
    "Short pass": "SHORT_PASS",
    "Medium Pass": "MEDIUM_PASS",
    "Long Pass": "LONG_PASS",
    "Play Action": "PLAY_ACTION",
    "Short Run": "INSIDE_RUN",
}

DEF_CAT = {
    "Zones": "ZONES",
    "Man": "MANS",
    "Zone Pressure": "ZONE_PRESSURE",
    "Man Pressure": "MAN_PRESSURE",
}


def _slug(formation: str, name: str) -> str:
    prefix = {
        "3-3 Stack": "33stk",
        "3-3 Stack 3-High": "33stk3h",
        "4-3": "43",
        "3-4": "34",
        "5-2": "52",
        "Nickel": "nickel",
        "Dime": "dime",
        "6-2": "62",
    }.get(formation, re.sub(r"[^a-z0-9]+", "_", formation.lower()).strip("_"))
    body = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return f"{prefix}_{body}"[:56]


def _parse_notes(notes: str | None) -> dict:
    if not notes:
        return {}
    text = str(notes).strip()
    meta: dict = {}
    if "concept:" in text.lower():
        m = re.search(r"concept:\s*([^|]+)", text, re.I)
        if m:
            meta["concept"] = m.group(1).strip()
    if "ball_carrier:" in text.lower():
        m = re.search(r"ball_carrier:\s*([^|]+)", text, re.I)
        if m:
            meta["ball_carrier"] = m.group(1).strip()
    elif "ball carrier" in text.lower() or "ball Carrier" in text:
        if "QB or RB" in text or "QB or TB" in text:
            meta["ball_carrier"] = "QB or RB"
        elif "QB is" in text or "QB Ball" in text or "QB ball" in text:
            meta["ball_carrier"] = "QB"
        elif "FB is" in text or "FB Ball" in text:
            meta["ball_carrier"] = "FB"
        elif "WR is" in text or "WR ball" in text:
            meta["ball_carrier"] = "WR"
        elif "RB is" in text or "RB Ball" in text or "Rb is" in text:
            meta["ball_carrier"] = "RB"
        elif "Ball carrier is QB" in text:
            meta["note"] = "Ball carrier is QB"
    elif text and not text.startswith("concept:"):
        meta["note"] = text
    return meta


def _read_offense_rows(ws) -> tuple[list[dict], list[dict]]:
    existing: list[dict] = []
    additions: list[dict] = []
    for r in range(2, 186):
        formation, name, pid, cat, _pbs, notes = [ws.cell(r, c).value for c in range(1, 7)]
        if not formation or not name:
            continue
        if str(formation).startswith("Offensive playbook") or str(formation).startswith("Total offensive"):
            break
        row = {
            "formation": str(formation).strip(),
            "name": str(name).strip(),
            "id": str(pid).strip() if pid else "",
            "category": str(cat).strip() if cat else "",
            "notes": str(notes).strip() if notes else "",
        }
        if row["id"]:
            existing.append(row)
    for r in range(196, ws.max_row + 1):
        formation, name, pid, cat, _pbs, notes = [ws.cell(r, c).value for c in range(1, 7)]
        if not formation or not name:
            continue
        row = {
            "formation": str(formation).strip(),
            "name": str(name).strip(),
            "id": str(pid).strip() if pid else _slug(str(formation), str(name)),
            "category": str(cat).strip() if cat else "",
            "notes": str(notes).strip() if notes else "",
        }
        additions.append(row)
    return existing, additions


def _read_defense_rows(ws) -> tuple[list[dict], list[dict]]:
    existing: list[dict] = []
    additions: list[dict] = []
    for r in range(2, ws.max_row + 1):
        formation, name, pid, cat, _pbs, notes = [ws.cell(r, c).value for c in range(1, 7)]
        if not formation or str(formation).startswith("Defensive playbook"):
            break
        if str(formation).startswith("Total defensive"):
            break
        row = {
            "formation": str(formation).strip(),
            "name": str(name).strip(),
            "id": str(pid).strip() if pid else "",
            "category": str(cat).strip() if cat else "",
            "notes": str(notes).strip() if notes else "",
        }
        if row["id"]:
            existing.append(row)
        elif row["name"]:
            row["id"] = _slug(row["formation"], row["name"])
            additions.append(row)
    return existing, additions


def _emit_play(side: str, row: dict) -> str:
    cat_map = OFF_CAT if side == "offense" else DEF_CAT
    cat_key = row["category"]
    if not cat_key and side == "offense" and "Quick Game" in row["notes"]:
        cat_key = "Short Pass"
    enum = cat_map.get(cat_key)
    if not enum:
        raise ValueError(f"Unknown category {cat_key!r} for {row}")
    cat_enum = "OffensivePlayCategory" if side == "offense" else "DefensivePlayCategory"
    cat_field = "offensive_category" if side == "offense" else "defensive_category"
    meta = _parse_notes(row["notes"])
    meta_repr = repr(meta) if meta else "{}"
    return (
        f"        Play(\n"
        f"            id={row['id']!r},\n"
        f"            name={row['name']!r},\n"
        f"            side={side!r},\n"
        f"            {cat_field}={cat_enum}.{enum},\n"
        f"            formation={row['formation']!r},\n"
        f"            metadata={meta_repr},\n"
        f"        ),"
    )


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"Missing xlsx: {xlsx}")

    from systems.formation_plays import _OFFENSIVE_FORMATIONS
    from systems.defensive_formations import _DEFENSIVE_FORMATIONS

    wb = load_workbook(xlsx, data_only=True)
    off_existing, off_add = _read_offense_rows(wb["Offense"])
    def_existing, def_add = _read_defense_rows(wb["Defense"])

    code_off = {(p.formation, p.id): p for f in _OFFENSIVE_FORMATIONS for p in _OFFENSIVE_FORMATIONS[f]}
    code_def = {(p.formation, p.id): p for f in _DEFENSIVE_FORMATIONS for p in _DEFENSIVE_FORMATIONS[f]}

    off_renames: dict[str, dict] = {}
    for row in off_existing:
        p = code_off.get((row["formation"], row["id"]))
        if not p:
            continue
        patch: dict = {}
        if p.name.strip() != row["name"]:
            patch["name"] = row["name"]
        new_meta = _parse_notes(row["notes"])
        if new_meta:
            if new_meta.get("note") == "Ball carrier is QB":
                patch["metadata"] = new_meta
            else:
                merged = dict(p.metadata or {})
                merged.update(new_meta)
                if merged != dict(p.metadata or {}):
                    patch["metadata"] = merged
        if patch:
            off_renames[row["id"]] = patch

    def_renames: dict[str, dict] = {}
    for row in def_existing:
        p = code_def.get((row["formation"], row["id"]))
        if not p:
            continue
        if p.name.strip() != row["name"]:
            def_renames[row["id"]] = {"name": row["name"]}

    # group additions
    off_by_form: dict[str, list[dict]] = {}
    for row in off_add:
        off_by_form.setdefault(row["formation"], []).append(row)
    def_by_form: dict[str, list[dict]] = {}
    for row in def_add:
        def_by_form.setdefault(row["formation"], []).append(row)

    lines: list[str] = [
        '"""',
        "Playbook overlay: renames and new plays from FND_Playbook_Reference.xlsx.",
        "Generated by scripts/build_playbook_overlay.py — do not edit by hand.",
        '"""',
        "from __future__ import annotations",
        "",
        "from typing import Dict, List",
        "",
        "from models.play import Play, OffensivePlayCategory, DefensivePlayCategory",
        "",
        f"OFFENSIVE_RENAMES: Dict[str, dict] = {repr(off_renames)}",
        "",
        f"DEFENSIVE_RENAMES: Dict[str, dict] = {repr(def_renames)}",
        "",
        "OFFENSIVE_EXTRA_FORMATIONS: Dict[str, List[Play]] = {",
    ]
    for formation in sorted(off_by_form):
        lines.append(f'    {formation!r}: [')
        for row in off_by_form[formation]:
            lines.append(_emit_play("offense", row))
        lines.append("    ],")
    lines.append("}")
    lines.extend(["", "DEFENSIVE_EXTRA_FORMATIONS: Dict[str, List[Play]] = {"])
    for formation in sorted(def_by_form):
        lines.append(f'    {formation!r}: [')
        for row in def_by_form[formation]:
            lines.append(_emit_play("defense", row))
        lines.append("    ],")
    lines.append("}")
    lines.extend(
        [
            "",
            "GOALLINE_FORMATION = \"Goalline\"",
            "WIDE_SLOT_FORMATION = \"Wide Slot\"",
            "",
            "def _apply_renames(plays: List[Play], renames: Dict[str, dict]) -> List[Play]:",
            "    out: List[Play] = []",
            "    for p in plays:",
            "        patch = renames.get(p.id)",
            "        if not patch:",
            "            out.append(p)",
            "            continue",
            "        kwargs = dict(",
            "            id=p.id,",
            "            name=patch.get(\"name\", p.name),",
            "            side=p.side,",
            "            offensive_category=p.offensive_category,",
            "            defensive_category=p.defensive_category,",
            "            formation=p.formation,",
            "            metadata=patch.get(\"metadata\", p.metadata),",
            "        )",
            "        out.append(Play(**kwargs))",
            "    return out",
            "",
            "def overlay_offensive_plays(formation: str, plays: List[Play]) -> List[Play]:",
            "    merged = _apply_renames(list(plays), OFFENSIVE_RENAMES)",
            "    extras = OFFENSIVE_EXTRA_FORMATIONS.get(formation, [])",
            "    if not extras:",
            "        return merged",
            "    seen = {p.id for p in merged}",
            "    for p in extras:",
            "        if p.id not in seen:",
            "            merged.append(p)",
            "            seen.add(p.id)",
            "    return merged",
            "",
            "def overlay_defensive_plays(formation: str, plays: List[Play]) -> List[Play]:",
            "    merged = _apply_renames(list(plays), DEFENSIVE_RENAMES)",
            "    extras = DEFENSIVE_EXTRA_FORMATIONS.get(formation, [])",
            "    if not extras:",
            "        return merged",
            "    seen = {p.id for p in merged}",
            "    for p in extras:",
            "        if p.id not in seen:",
            "            merged.append(p)",
            "            seen.add(p.id)",
            "    return merged",
            "",
            "def extra_offensive_formations() -> List[str]:",
            "    return list(OFFENSIVE_EXTRA_FORMATIONS.keys())",
            "",
        ]
    )

    OUT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_PATH}")
    print(f"  offense renames: {len(off_renames)}")
    print(f"  offense new plays: {sum(len(v) for v in off_by_form.values())}")
    print(f"  defense renames: {len(def_renames)}")
    print(f"  defense new plays: {sum(len(v) for v in def_by_form.values())}")


if __name__ == "__main__":
    main()
