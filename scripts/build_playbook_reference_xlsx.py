#!/usr/bin/env python3
"""Generate playbook reference Excel: all offensive/defensive formations, plays, and playbooks.

Requires: pip install openpyxl
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATHS = [
    ROOT / "data" / "templates" / "FND_Playbook_Reference.xlsx",
    ROOT / "frontend" / "public" / "templates" / "FND_Playbook_Reference.xlsx",
]

from systems.defensive_formations import list_defensive_formations, get_defensive_formation_plays
from systems.formation_plays import list_formations, get_formation_plays
from systems.playbook_system import (
    DEFENSIVE_PLAYBOOK_FORMATIONS,
    OFFENSIVE_PLAYBOOK_FORMATIONS,
    OFFENSIVE_PLAYBOOK_VALUES,
    DEFENSIVE_PLAYBOOK_KEYS,
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F3F4F6")


def _playbooks_for_formation(
    formation: str,
    playbook_map: Dict[str, List[str]],
) -> List[str]:
  """Return playbook names that include this formation, in canonical order."""
  found: List[str] = []
  for playbook, formations in playbook_map.items():
    if formation in formations:
      found.append(playbook)
  return found


def _metadata_note(metadata: dict) -> str:
  if not metadata:
    return ""
  parts: List[str] = []
  for key in ("note", "concept", "ball_carrier"):
    val = metadata.get(key)
    if val:
      parts.append(f"{key}: {val}")
  for key, val in metadata.items():
    if key not in ("note", "concept", "ball_carrier") and val:
      parts.append(f"{key}: {val}")
  return " | ".join(parts)


def _style_sheet(ws, widths: List[int]) -> None:
  ws.freeze_panes = "A2"
  for col_idx, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width
  for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  for row_idx in range(2, ws.max_row + 1):
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for cell in ws[row_idx]:
      if fill:
        cell.fill = fill
      cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_offense_rows() -> List[List[str]]:
  rows: List[List[str]] = []
  for formation in list_formations():
    playbooks = _playbooks_for_formation(formation, OFFENSIVE_PLAYBOOK_FORMATIONS)
    for play in get_formation_plays(formation):
      category = play.offensive_category.value if play.offensive_category else ""
      rows.append(
        [
          formation,
          play.name,
          play.id,
          category,
          ", ".join(playbooks),
          _metadata_note(play.metadata),
        ]
      )
  rows.sort(key=lambda r: (r[0], r[3], r[1]))
  return rows


def _build_defense_rows() -> List[List[str]]:
  rows: List[List[str]] = []
  for formation in list_defensive_formations():
    playbooks = _playbooks_for_formation(formation, DEFENSIVE_PLAYBOOK_FORMATIONS)
    for play in get_defensive_formation_plays(formation):
      category = play.defensive_category.value if play.defensive_category else ""
      rows.append(
        [
          formation,
          play.name,
          play.id,
          category,
          ", ".join(playbooks),
          _metadata_note(play.metadata),
        ]
      )
  rows.sort(key=lambda r: (r[0], r[3], r[1]))
  return rows


def _add_summary_block(ws, start_row: int, title: str, playbook_map: Dict[str, List[str]], keys: tuple) -> int:
  ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
  row = start_row + 1
  ws.cell(row=row, column=1, value="Playbook").font = Font(bold=True)
  ws.cell(row=row, column=2, value="Formations included").font = Font(bold=True)
  row += 1
  for key in keys:
    formations = ", ".join(playbook_map.get(key, []))
    ws.cell(row=row, column=1, value=key)
    ws.cell(row=row, column=2, value=formations)
    row += 1
  return row + 1


def build_workbook() -> Workbook:
  wb = Workbook()

  off_headers = ["Formation", "Play Name", "Play ID", "Category", "Playbook(s)", "Notes"]
  def_headers = off_headers

  # Offense sheet
  ws_off = wb.active
  ws_off.title = "Offense"
  ws_off.append(off_headers)
  off_rows = _build_offense_rows()
  off_play_count = len(off_rows)
  for row in off_rows:
    ws_off.append(row)
  _style_sheet(ws_off, [16, 22, 28, 14, 36, 48])

  summary_row = ws_off.max_row + 3
  next_row = _add_summary_block(
    ws_off,
    summary_row,
    "Offensive playbook → formations",
    OFFENSIVE_PLAYBOOK_FORMATIONS,
    OFFENSIVE_PLAYBOOK_VALUES,
  )
  ws_off.cell(row=next_row, column=1, value=f"Total offensive plays listed: {off_play_count}")

  # Defense sheet
  ws_def = wb.create_sheet("Defense")
  ws_def.append(def_headers)
  def_rows = _build_defense_rows()
  def_play_count = len(def_rows)
  for row in def_rows:
    ws_def.append(row)
  _style_sheet(ws_def, [18, 22, 28, 16, 36, 56])

  summary_row = ws_def.max_row + 3
  next_row = _add_summary_block(
    ws_def,
    summary_row,
    "Defensive playbook → formations",
    DEFENSIVE_PLAYBOOK_FORMATIONS,
    DEFENSIVE_PLAYBOOK_KEYS,
  )
  ws_def.cell(row=next_row, column=1, value=f"Total defensive plays listed: {def_play_count}")

  return wb


def main() -> None:
  wb = build_workbook()
  for path in OUT_PATHS:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


if __name__ == "__main__":
  main()
